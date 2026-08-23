#!/usr/bin/env python3
"""Apply title, meta, and H1 updates from NutriThrive_Exact_Titles_Metas_55_Blogs.xlsx."""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "NutriThrive_Exact_Titles_Metas_55_Blogs.xlsx"
BLOG_DIR = ROOT / "blog"
MODIFIED = "2026-08-23T00:00:00+10:00"
MODIFIED_SHORT = "2026-08-23"


def html_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def resolve_blog_path(slug: str) -> Path | None:
    for name in (f"{slug}.html", slug):
        path = BLOG_DIR / name
        if path.is_file():
            return path
    return None


def load_updates() -> dict[str, dict[str, str]]:
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    updates: dict[str, dict[str, str]] = {}
    for row in rows:
        url, _, _, title, _, meta, _, h1 = row
        if not url or not title or not meta or not h1:
            continue
        slug = str(url).replace("https://nutrithrive.com.au/blog/", "").rstrip("/")
        path = resolve_blog_path(slug)
        if not path:
            continue
        updates[path.name] = {
            "title": str(title),
            "meta": str(meta),
            "h1": str(h1),
        }
    return updates


def update_blog_file(path: Path, data: dict[str, str]) -> None:
    content = path.read_text(encoding="utf-8")
    title, meta, h1 = data["title"], data["meta"], data["h1"]
    esc_title, esc_meta, esc_h1 = map(html_escape, (title, meta, h1))

    content = re.sub(r"<title>[^<]*</title>", f"<title>{esc_title}</title>", content, count=1)
    content = re.sub(
        r'<meta name="description" content="[^"]*"',
        f'<meta name="description" content="{esc_meta}"',
        content,
        count=1,
    )

    content = re.sub(
        r'<meta property="og:title" content="[^"]*"',
        f'<meta property="og:title" content="{esc_title}"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta content="[^"]*" property="og:title"',
        f'<meta content="{esc_title}" property="og:title"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta content="[^"]*" property="twitter:title"',
        f'<meta content="{esc_title}" property="twitter:title"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta name="twitter:title" content="[^"]*"',
        f'<meta name="twitter:title" content="{esc_title}"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta property="og:description" content="[^"]*"',
        f'<meta property="og:description" content="{esc_meta}"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta content="[^"]*" property="og:description"',
        f'<meta content="{esc_meta}" property="og:description"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta content="[^"]*" property="twitter:description"',
        f'<meta content="{esc_meta}" property="twitter:description"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta name="twitter:description" content="[^"]*"',
        f'<meta name="twitter:description" content="{esc_meta}"',
        content,
        count=1,
    )

    def replace_first_h1(match: re.Match) -> str:
        return f"{match.group(1)}{esc_h1}{match.group(3)}"

    content = re.sub(
        r'(<h1 class="font-display[^"]*"[^>]*>)([^<]*)(</h1>)',
        replace_first_h1,
        content,
        count=1,
    )
    if f">{esc_h1}</h1>" not in content:
        content = re.sub(
            r"(<h1[^>]*>)([^<]*)(</h1>)",
            replace_first_h1,
            content,
            count=1,
        )

    def patch_ld_json(match: re.Match) -> str:
        try:
            obj = json.loads(match.group(1))
        except json.JSONDecodeError:
            return match.group(0)
        if obj.get("@type") in ("BlogPosting", "Article"):
            obj["headline"] = title
            obj["description"] = meta
            if "dateModified" in obj:
                obj["dateModified"] = MODIFIED_SHORT
            return (
                '<script type="application/ld+json">'
                + json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
                + "</script>"
            )
        if obj.get("@type") == "BreadcrumbList":
            items = obj.get("itemListElement", [])
            if items:
                items[-1]["name"] = title
            return (
                '<script type="application/ld+json">'
                + json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
                + "</script>"
            )
        return match.group(0)

    content = re.sub(
        r'<script type="application/ld\+json">(\{.*?\})</script>',
        patch_ld_json,
        content,
        flags=re.DOTALL,
    )

    content = re.sub(
        r'<meta property="article:modified_time" content="[^"]*"',
        f'<meta property="article:modified_time" content="{MODIFIED}"',
        content,
        count=1,
    )
    content = re.sub(
        r'<meta content="2026-[^"]*" name="last-modified"',
        f'<meta content="{MODIFIED_SHORT}" name="last-modified"',
        content,
        count=1,
    )

    path.write_text(content, encoding="utf-8")


def main() -> None:
    updates = load_updates()
    missing_slugs = []
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    for row in rows:
        url = row[0]
        if not url:
            continue
        slug = str(url).replace("https://nutrithrive.com.au/blog/", "").rstrip("/")
        if not resolve_blog_path(slug):
            missing_slugs.append(slug)

    print(f"Updating {len(updates)} blog posts...")
    for filename, data in sorted(updates.items()):
        path = BLOG_DIR / filename
        update_blog_file(path, data)
        print(f"  {filename}")

    if missing_slugs:
        print(f"\nSkipped {len(missing_slugs)} posts (file not found):")
        for slug in missing_slugs:
            print(f"  {slug}")


if __name__ == "__main__":
    main()
